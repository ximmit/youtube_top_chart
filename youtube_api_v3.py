import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from apiclient.discovery import build
from openpyxl import Workbook
api_key="AIzaSyCzk1hhI8kg0HCL1fMHJTQQZf10bgkzDm8"
youtube=build('youtube','v3', developerKey=api_key)
type(youtube)
req=youtube.search().list(part="snippet",order="viewCount",type='video',maxResults=50)
#req=youtube.videos().list(part="snippet,contentDetails,statistics",chart="mostPopular",maxResults=50)
res=req.execute()
#print(res['items'][0])
i=1
wb = Workbook()
ws = wb.active
for it in res['items']:
 print(it['snippet']['title'])
 ws["A"+str(i)].value = it['snippet']['title']
 ws.cell(row=i, column=2).value ="https: // www.youtube.com / watch?v ="+it['id']['videoId']

 i=i+1

wb.save("youtube_chart.xlsx")
 #print(it['id']['videoId'])
